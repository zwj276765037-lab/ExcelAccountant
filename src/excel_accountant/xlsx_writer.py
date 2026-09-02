from __future__ import annotations

import hashlib
import os
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from .decimal_codec import format_decimal
from .models import (
    ExactSolution,
    OutputArtifact,
    TargetAmount,
    WorkbookPreview,
)
from .output_verifier import (
    OutputVerificationError,
    target_color,
    validate_solution_structure,
    verify_output_workbook,
)


class UnsafeWorkbookOutputError(ValueError):
    """Raised when a workbook is intentionally blocked from output."""


def file_sha256(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_exact_solution(
    preview: WorkbookPreview,
    targets: tuple[TargetAmount, ...],
    solution: ExactSolution,
    output_directory: str | Path,
    *,
    scheme_number: int,
) -> OutputArtifact:
    if not preview.safety.safe_to_write:
        reasons = "；".join(preview.safety.reasons)
        raise UnsafeWorkbookOutputError(f"工作簿禁止输出：{reasons}")
    if scheme_number < 1:
        raise ValueError("方案编号必须从 1 开始")
    validate_solution_structure(preview, targets, solution)

    source_path = preview.path
    original_hash = file_sha256(source_path)
    output_dir = Path(output_directory)
    output_dir.mkdir(parents=True, exist_ok=True)
    final_path = _available_path(
        output_dir / f"{source_path.stem}_方案{scheme_number:03d}.xlsx"
    )
    file_handle, temporary_name = tempfile.mkstemp(
        prefix=".excel_accountant_",
        suffix=".xlsx",
        dir=output_dir,
    )
    os.close(file_handle)
    temporary_path = Path(temporary_name)

    try:
        shutil.copy2(source_path, temporary_path)
        workbook = load_workbook(
            temporary_path,
            read_only=False,
            data_only=False,
            keep_links=True,
        )
        try:
            source_sheet = workbook[preview.sheet]
            result_sheet_name = _unique_sheet_name(workbook.sheetnames, "凑数结果")
            result_sheet = workbook.create_sheet(result_sheet_name)
            _write_audit_header(result_sheet, scheme_number)
            for row_index, assignment in enumerate(solution.assignments, start=5):
                target = targets[assignment.target_index]
                color = target_color(assignment.target_index)
                fill = PatternFill(fill_type="solid", fgColor=f"FF{color}")
                source_cells = [preview.cells[index] for index in assignment.cell_indices]
                for source_cell in source_cells:
                    source_sheet[source_cell.address].fill = fill
                exact_sum = sum(
                    (cell.amount for cell in source_cells),
                    start=target.amount * 0,
                )
                values = ", ".join(format_decimal(cell.amount) for cell in source_cells)
                addresses = ", ".join(cell.address for cell in source_cells)
                row_values = (
                    scheme_number,
                    target.identifier,
                    target.raw_value,
                    color,
                    values,
                    addresses,
                    len(source_cells),
                    format_decimal(exact_sum),
                    "通过",
                )
                for column, value in enumerate(row_values, start=1):
                    result_sheet.cell(row_index, column, value)
                result_sheet.cell(row_index, 4).fill = fill
            _format_audit_sheet(result_sheet)
            workbook.save(temporary_path)
        finally:
            workbook.close()

        verify_output_workbook(
            temporary_path,
            preview,
            targets,
            solution,
            result_sheet_name,
        )
        if file_sha256(source_path) != original_hash:
            raise OutputVerificationError("原始工作簿哈希发生变化")
        os.replace(temporary_path, final_path)
        return OutputArtifact(final_path, result_sheet_name, scheme_number)
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise


def _write_audit_header(worksheet, scheme_number: int) -> None:
    worksheet["A1"] = f"ExcelAccountant 凑数结果 - 方案 {scheme_number:03d}"
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A2"] = "同一文件中每个源单元格最多使用一次。"
    headers = (
        "方案编号",
        "目标编号",
        "目标金额",
        "颜色",
        "组成金额",
        "单元格地址",
        "使用数量",
        "精确合计",
        "复核状态",
    )
    for column, value in enumerate(headers, start=1):
        cell = worksheet.cell(4, column, value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="FFD9EAF7")


def _format_audit_sheet(worksheet) -> None:
    widths = (12, 18, 20, 12, 45, 45, 12, 20, 12)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
    worksheet.freeze_panes = "A5"


def _unique_sheet_name(existing: list[str], base: str) -> str:
    if base not in existing:
        return base
    index = 1
    while f"{base}_{index}" in existing:
        index += 1
    return f"{base}_{index}"


def _available_path(path: Path) -> Path:
    if not path.exists():
        return path
    index = 1
    while True:
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
        index += 1
