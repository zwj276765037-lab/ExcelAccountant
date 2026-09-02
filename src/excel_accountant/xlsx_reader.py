from __future__ import annotations

import posixpath
import re
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries

from .decimal_codec import AmountParseError, parse_amount, parse_storage_decimal
from .models import MoneyCell, SkippedCell, WorkbookPreview
from .workbook_safety import WorkbookSafetyError, inspect_workbook_safety

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


class WorkbookReadError(ValueError):
    """Raised when an XLSX input or requested range cannot be read."""


@dataclass(frozen=True, slots=True)
class ColumnRange:
    column: int
    start_row: int = 1
    end_row: int | None = None


@dataclass(frozen=True, slots=True)
class RawCell:
    address: str
    cell_type: str
    raw_value: str
    has_formula: bool


def parse_column_range(text: str) -> ColumnRange:
    value = text.strip().upper()
    if not value:
        raise WorkbookReadError("金额列不能为空")
    if value.isdigit():
        column = int(value)
        if not 1 <= column <= 16_384:
            raise WorkbookReadError("列序号必须在 1 至 16384 之间")
        return ColumnRange(column)
    if re.fullmatch(r"[A-Z]{1,3}", value):
        try:
            return ColumnRange(column_index_from_string(value))
        except ValueError as exc:
            raise WorkbookReadError(f"无效列字母：{text}") from exc
    try:
        min_col, min_row, max_col, max_row = range_boundaries(value)
    except ValueError as exc:
        raise WorkbookReadError(f"无效单列范围：{text}") from exc
    if min_col != max_col:
        raise WorkbookReadError("第一版只支持单列范围")
    if min_row == 1 and max_row == 1_048_576:
        return ColumnRange(min_col)
    return ColumnRange(min_col, min_row, max_row)


def list_sheet_names(path: str | Path) -> tuple[str, ...]:
    try:
        with ZipFile(path) as archive:
            mappings = _sheet_paths(archive)
    except (BadZipFile, OSError, ET.ParseError, KeyError) as exc:
        raise WorkbookReadError(f"无法读取工作表列表：{exc}") from exc
    return tuple(mappings)


def read_workbook_preview(
    path: str | Path,
    sheet: str,
    range_text: str,
) -> WorkbookPreview:
    workbook_path = Path(path)
    requested = parse_column_range(range_text)
    try:
        safety = inspect_workbook_safety(workbook_path)
    except WorkbookSafetyError as exc:
        raise WorkbookReadError(str(exc)) from exc

    try:
        workbook = load_workbook(
            workbook_path,
            read_only=False,
            data_only=False,
            keep_links=True,
        )
    except Exception as exc:
        raise WorkbookReadError(f"无法打开工作簿：{exc}") from exc

    try:
        if sheet not in workbook.sheetnames:
            raise WorkbookReadError(f"工作表不存在：{sheet}")
        worksheet = workbook[sheet]
        with ZipFile(workbook_path) as archive:
            sheet_paths = _sheet_paths(archive)
            if sheet not in sheet_paths:
                raise WorkbookReadError(f"无法定位工作表 XML：{sheet}")
            shared_strings = _shared_strings(archive)
            raw_cells = _read_raw_cells(
                archive,
                sheet_paths[sheet],
                requested,
                shared_strings,
            )

        cells: list[MoneyCell] = []
        skipped: list[SkippedCell] = []
        formula_count = 0
        zero_count = 0
        hidden_count = 0

        for raw_cell in raw_cells:
            cell = worksheet[raw_cell.address]
            if raw_cell.has_formula:
                formula_count += 1
                skipped.append(
                    SkippedCell(sheet, raw_cell.address, raw_cell.raw_value, "公式单元格")
                )
                continue
            if cell.is_date:
                skipped.append(
                    SkippedCell(sheet, raw_cell.address, raw_cell.raw_value, "日期单元格")
                )
                continue
            if raw_cell.cell_type == "b":
                skipped.append(
                    SkippedCell(sheet, raw_cell.address, raw_cell.raw_value, "布尔值")
                )
                continue
            if raw_cell.cell_type == "e":
                skipped.append(
                    SkippedCell(sheet, raw_cell.address, raw_cell.raw_value, "错误值")
                )
                continue

            try:
                if raw_cell.cell_type in {"s", "inlineStr", "str"}:
                    amount = parse_amount(raw_cell.raw_value)
                    source_type = "text"
                else:
                    amount = parse_storage_decimal(raw_cell.raw_value)
                    source_type = "number"
            except AmountParseError:
                reason = (
                    "表头"
                    if cell.row == requested.start_row and requested.start_row == 1
                    else "非严格金额文本"
                )
                skipped.append(
                    SkippedCell(sheet, raw_cell.address, raw_cell.raw_value, reason)
                )
                continue

            if amount == 0:
                zero_count += 1
                skipped.append(
                    SkippedCell(sheet, raw_cell.address, raw_cell.raw_value, "零值")
                )
                continue

            hidden = bool(worksheet.row_dimensions[cell.row].hidden)
            if hidden:
                hidden_count += 1
            cells.append(
                MoneyCell(
                    sheet=sheet,
                    address=raw_cell.address,
                    row=cell.row,
                    column=cell.column,
                    raw_value=raw_cell.raw_value,
                    amount=amount,
                    source_type=source_type,
                    hidden=hidden,
                )
            )

        last_row = max((cell.address for cell in raw_cells), default=None)
        normalized_range = _normalized_range(requested, raw_cells)
        return WorkbookPreview(
            path=workbook_path,
            sheet=sheet,
            range_text=normalized_range,
            cells=tuple(cells),
            skipped=tuple(skipped),
            formula_count=formula_count,
            zero_count=zero_count,
            hidden_count=hidden_count,
            safety=safety,
        )
    except (BadZipFile, ET.ParseError, KeyError, IndexError) as exc:
        raise WorkbookReadError(f"工作簿 XML 解析失败：{exc}") from exc
    finally:
        workbook.close()


def _sheet_paths(archive: ZipFile) -> dict[str, str]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    relationships_root = ET.fromstring(
        archive.read("xl/_rels/workbook.xml.rels")
    )
    relationship_targets = {
        node.attrib["Id"]: node.attrib["Target"]
        for node in relationships_root.findall(f"{{{PKG_REL_NS}}}Relationship")
    }
    result: dict[str, str] = {}
    for sheet in workbook_root.findall(f".//{{{MAIN_NS}}}sheet"):
        relationship_id = sheet.attrib[f"{{{DOC_REL_NS}}}id"]
        target = relationship_targets[relationship_id]
        if target.startswith("/"):
            path = target.lstrip("/")
        else:
            path = posixpath.normpath(posixpath.join("xl", target))
        result[sheet.attrib["name"]] = path
    return result


def _shared_strings(archive: ZipFile) -> tuple[str, ...]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return ()
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    values: list[str] = []
    for item in root.findall(f"{{{MAIN_NS}}}si"):
        text = "".join(
            node.text or "" for node in item.iter(f"{{{MAIN_NS}}}t")
        )
        values.append(text)
    return tuple(values)


def _read_raw_cells(
    archive: ZipFile,
    sheet_path: str,
    requested: ColumnRange,
    shared_strings: tuple[str, ...],
) -> tuple[RawCell, ...]:
    root = ET.fromstring(archive.read(sheet_path))
    result: list[RawCell] = []
    for node in root.iter(f"{{{MAIN_NS}}}c"):
        address = node.attrib.get("r")
        if not address:
            continue
        match = re.fullmatch(r"([A-Z]+)(\d+)", address.upper())
        if not match:
            continue
        column = column_index_from_string(match.group(1))
        row = int(match.group(2))
        if column != requested.column or row < requested.start_row:
            continue
        if requested.end_row is not None and row > requested.end_row:
            continue
        cell_type = node.attrib.get("t", "n")
        formula = node.find(f"{{{MAIN_NS}}}f") is not None
        value_node = node.find(f"{{{MAIN_NS}}}v")
        if cell_type == "inlineStr":
            raw_value = "".join(
                text_node.text or ""
                for text_node in node.iter(f"{{{MAIN_NS}}}t")
            )
        elif value_node is None or value_node.text is None:
            raw_value = ""
        elif cell_type == "s":
            raw_value = shared_strings[int(value_node.text)]
        else:
            raw_value = value_node.text
        if raw_value == "" and not formula:
            continue
        result.append(RawCell(address, cell_type, raw_value, formula))
    result.sort(key=lambda cell: int(re.search(r"\d+$", cell.address).group()))
    return tuple(result)


def _normalized_range(
    requested: ColumnRange, raw_cells: tuple[RawCell, ...]
) -> str:
    letter = get_column_letter(requested.column)
    if requested.end_row is not None:
        end_row = requested.end_row
    else:
        end_row = max(
            (int(re.search(r"\d+$", cell.address).group()) for cell in raw_cells),
            default=requested.start_row,
        )
    return f"{letter}{requested.start_row}:{letter}{end_row}"
