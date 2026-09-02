from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .decimal_codec import format_decimal
from .models import (
    ApproximateSolution,
    ExactSolution,
    MoneyCell,
    ScaledProblem,
    TargetAmount,
    WorkbookPreview,
)
from .xlsx_reader import read_workbook_preview


class OutputVerificationError(ValueError):
    """Raised when a generated workbook fails an independent audit."""


def target_color(target_index: int) -> str:
    palette = (
        "FFC7CE",
        "C6EFCE",
        "FFEB9C",
        "BDD7EE",
        "E4DFEC",
        "F4B183",
        "A9D18E",
        "9DC3E6",
        "D5A6BD",
        "B4C6E7",
        "FFD966",
        "A5A5A5",
        "70AD47",
        "5B9BD5",
        "ED7D31",
        "A5A5FF",
        "66CCCC",
        "CC99FF",
        "FF9999",
        "99CC00",
    )
    return palette[target_index % len(palette)]


def validate_solution_structure(
    preview: WorkbookPreview,
    targets: tuple[TargetAmount, ...],
    solution: ExactSolution,
) -> None:
    if len(solution.assignments) != len(targets):
        raise OutputVerificationError("方案目标数量与输入目标数量不一致")
    used: set[int] = set()
    for assignment in solution.assignments:
        if not 0 <= assignment.target_index < len(targets):
            raise OutputVerificationError("方案包含无效目标索引")
        if not assignment.cell_indices:
            raise OutputVerificationError("精确目标组合不能为空")
        for cell_index in assignment.cell_indices:
            if not 0 <= cell_index < len(preview.cells):
                raise OutputVerificationError("方案包含无效源单元格索引")
            if cell_index in used:
                raise OutputVerificationError("同一源单元格在方案中被重复使用")
            used.add(cell_index)
        total = sum(
            (preview.cells[index].amount for index in assignment.cell_indices),
            start=targets[assignment.target_index].amount * 0,
        )
        if total != targets[assignment.target_index].amount:
            raise OutputVerificationError(
                f"目标 {targets[assignment.target_index].identifier} 的组合合计不精确"
            )


def verify_output_workbook(
    output_path: str | Path,
    source_preview: WorkbookPreview,
    targets: tuple[TargetAmount, ...],
    solution: ExactSolution,
    result_sheet: str,
) -> None:
    validate_solution_structure(source_preview, targets, solution)
    output_preview = read_workbook_preview(
        output_path,
        source_preview.sheet,
        source_preview.range_text,
    )
    output_values = {cell.address: cell.amount for cell in output_preview.cells}
    workbook = load_workbook(output_path, read_only=False, data_only=False)
    try:
        if result_sheet not in workbook.sheetnames:
            raise OutputVerificationError("输出缺少凑数结果工作表")
        source_sheet = workbook[source_preview.sheet]
        audit_sheet = workbook[result_sheet]
        seen_addresses: set[str] = set()
        for row_offset, assignment in enumerate(solution.assignments, start=5):
            target = targets[assignment.target_index]
            addresses: list[str] = []
            actual = target.amount * 0
            expected_color = target_color(assignment.target_index)
            for cell_index in assignment.cell_indices:
                source_cell = source_preview.cells[cell_index]
                if source_cell.address in seen_addresses:
                    raise OutputVerificationError("输出中存在重复源地址")
                seen_addresses.add(source_cell.address)
                if source_cell.address not in output_values:
                    raise OutputVerificationError(
                        f"输出缺少源金额：{source_cell.address}"
                    )
                actual += output_values[source_cell.address]
                addresses.append(source_cell.address)
                fill_rgb = source_sheet[source_cell.address].fill.fgColor.rgb
                if not fill_rgb or not str(fill_rgb).upper().endswith(expected_color):
                    raise OutputVerificationError(
                        f"单元格颜色验证失败：{source_cell.address}"
                    )
            if actual != target.amount:
                raise OutputVerificationError(
                    f"输出重读后目标 {target.identifier} 合计不精确"
                )
            if audit_sheet.cell(row_offset, 2).value != target.identifier:
                raise OutputVerificationError("结果表目标编号不一致")
            if audit_sheet.cell(row_offset, 6).value != ", ".join(addresses):
                raise OutputVerificationError("结果表源地址不一致")
            if audit_sheet.cell(row_offset, 9).value != "通过":
                raise OutputVerificationError("结果表复核状态不正确")
    finally:
        workbook.close()


def validate_approximate_solution_structure(
    preview: WorkbookPreview,
    targets: tuple[TargetAmount, ...],
    problem: ScaledProblem,
    solution: ApproximateSolution,
) -> None:
    if len(solution.assignments) != len(targets):
        raise OutputVerificationError("近似方案目标数量与输入不一致")
    used: set[int] = set()
    seen_targets: set[int] = set()
    for assignment in solution.assignments:
        if not 0 <= assignment.target_index < len(targets):
            raise OutputVerificationError("近似方案包含无效目标索引")
        if assignment.target_index in seen_targets:
            raise OutputVerificationError("近似方案中目标被重复分配")
        seen_targets.add(assignment.target_index)
        target = targets[assignment.target_index]
        for cell_index in assignment.cell_indices:
            if not 0 <= cell_index < len(preview.cells):
                raise OutputVerificationError("近似方案包含无效源单元格索引")
            if cell_index in used:
                raise OutputVerificationError("同一源单元格在近似方案中被重复使用")
            used.add(cell_index)
        actual = sum(
            (preview.cells[index].amount for index in assignment.cell_indices),
            start=target.amount * 0,
        )
        if actual != problem.restore(assignment.actual):
            raise OutputVerificationError("近似方案的实际合计与源单元格不一致")
        if actual - target.amount != problem.restore(assignment.difference):
            raise OutputVerificationError("近似方案的差额记录不正确")


def verify_approximate_output_workbook(
    output_path: str | Path,
    source_preview: WorkbookPreview,
    targets: tuple[TargetAmount, ...],
    problem: ScaledProblem,
    solution: ApproximateSolution,
    result_sheet: str,
) -> None:
    validate_approximate_solution_structure(source_preview, targets, problem, solution)
    output_preview = read_workbook_preview(
        output_path,
        source_preview.sheet,
        source_preview.range_text,
    )
    output_values = {cell.address: cell.amount for cell in output_preview.cells}
    workbook = load_workbook(output_path, read_only=False, data_only=False)
    try:
        if result_sheet not in workbook.sheetnames:
            raise OutputVerificationError("输出缺少近似凑数结果工作表")
        source_sheet = workbook[source_preview.sheet]
        audit_sheet = workbook[result_sheet]
        seen_addresses: set[str] = set()
        for row_offset, assignment in enumerate(solution.assignments, start=6):
            target = targets[assignment.target_index]
            addresses: list[str] = []
            actual = target.amount * 0
            expected_color = target_color(assignment.target_index)
            for cell_index in assignment.cell_indices:
                source_cell = source_preview.cells[cell_index]
                if source_cell.address in seen_addresses:
                    raise OutputVerificationError("近似输出中存在重复源地址")
                seen_addresses.add(source_cell.address)
                if source_cell.address not in output_values:
                    raise OutputVerificationError(
                        f"近似输出缺少源金额：{source_cell.address}"
                    )
                actual += output_values[source_cell.address]
                addresses.append(source_cell.address)
                fill_rgb = source_sheet[source_cell.address].fill.fgColor.rgb
                if not fill_rgb or not str(fill_rgb).upper().endswith(expected_color):
                    raise OutputVerificationError(
                        f"近似输出单元格颜色验证失败：{source_cell.address}"
                    )
            difference = actual - target.amount
            if actual != problem.restore(assignment.actual):
                raise OutputVerificationError("近似输出重读后实际合计不一致")
            if difference != problem.restore(assignment.difference):
                raise OutputVerificationError("近似输出重读后差额不一致")
            if audit_sheet.cell(row_offset, 2).value != target.identifier:
                raise OutputVerificationError("近似结果表目标编号不一致")
            if audit_sheet.cell(row_offset, 6).value != ", ".join(addresses):
                raise OutputVerificationError("近似结果表源地址不一致")
            if audit_sheet.cell(row_offset, 8).value != format_decimal(actual):
                raise OutputVerificationError("近似结果表实际合计不一致")
            if audit_sheet.cell(row_offset, 9).value != format_decimal(difference):
                raise OutputVerificationError("近似结果表差额不一致")
            status = str(audit_sheet.cell(row_offset, 11).value or "")
            if "近似" not in status:
                raise OutputVerificationError("近似结果表缺少风险标识")
    finally:
        workbook.close()
