from __future__ import annotations

from decimal import Decimal
from copy import copy
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from excel_accountant.models import (
    ExactSolution,
    TargetAmount,
    TargetAssignment,
    WorkbookSafetyReport,
)
from excel_accountant.output_verifier import (
    OutputVerificationError,
    verify_output_workbook,
)
from excel_accountant.xlsx_reader import read_workbook_preview
from excel_accountant.xlsx_writer import (
    UnsafeWorkbookOutputError,
    file_sha256,
    write_exact_solution,
)


def _workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "流水"
    sheet["E1"] = "交易金额"
    sheet["E2"] = 1.2
    sheet["E3"] = 2.3
    sheet["E4"] = 3.5
    sheet["A2"] = "项目A"
    sheet["A3"] = "项目B"
    sheet["A4"] = "项目C"
    workbook.save(path)


def _targets() -> tuple[TargetAmount, ...]:
    return (
        TargetAmount("目标1", "3.5", Decimal("3.5")),
        TargetAmount("目标2", "3.5", Decimal("3.5")),
    )


def _solution() -> ExactSolution:
    return ExactSolution(
        (
            TargetAssignment(0, (0, 1)),
            TargetAssignment(1, (2,)),
        )
    )


def test_write_and_verify_exact_solution(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _workbook(source)
    before_hash = file_sha256(source)
    preview = read_workbook_preview(source, "流水", "E")

    artifact = write_exact_solution(
        preview,
        _targets(),
        _solution(),
        tmp_path / "output",
        scheme_number=1,
    )

    assert artifact.path.exists()
    assert artifact.path != source
    assert file_sha256(source) == before_hash
    verify_output_workbook(
        artifact.path,
        preview,
        _targets(),
        _solution(),
        artifact.result_sheet,
    )
    workbook = load_workbook(artifact.path)
    try:
        assert artifact.result_sheet in workbook.sheetnames
        assert workbook["流水"]["E2"].fill.fgColor.rgb != workbook["流水"]["E4"].fill.fgColor.rgb
    finally:
        workbook.close()


def test_writer_never_overwrites_existing_output(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _workbook(source)
    preview = read_workbook_preview(source, "流水", "E")
    first = write_exact_solution(
        preview, _targets(), _solution(), tmp_path / "output", scheme_number=1
    )
    second = write_exact_solution(
        preview, _targets(), _solution(), tmp_path / "output", scheme_number=1
    )
    assert first.path != second.path
    assert first.path.exists() and second.path.exists()


def test_unsafe_workbook_cannot_be_written(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _workbook(source)
    preview = read_workbook_preview(source, "流水", "E")
    unsafe_preview = preview.__class__(
        path=preview.path,
        sheet=preview.sheet,
        range_text=preview.range_text,
        cells=preview.cells,
        skipped=preview.skipped,
        formula_count=preview.formula_count,
        zero_count=preview.zero_count,
        hidden_count=preview.hidden_count,
        safety=WorkbookSafetyReport(False, ("测试不安全结构",)),
    )
    with pytest.raises(UnsafeWorkbookOutputError):
        write_exact_solution(
            unsafe_preview,
            _targets(),
            _solution(),
            tmp_path / "output",
            scheme_number=1,
        )


def test_duplicate_source_index_is_rejected(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _workbook(source)
    preview = read_workbook_preview(source, "流水", "E")
    invalid = ExactSolution(
        (
            TargetAssignment(0, (0, 1)),
            TargetAssignment(1, (0, 1)),
        )
    )
    with pytest.raises(OutputVerificationError, match="重复使用"):
        write_exact_solution(
            preview,
            _targets(),
            invalid,
            tmp_path / "output",
            scheme_number=1,
        )


def test_tampered_fill_fails_independent_verification(tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    _workbook(source)
    preview = read_workbook_preview(source, "流水", "E")
    artifact = write_exact_solution(
        preview, _targets(), _solution(), tmp_path / "output", scheme_number=1
    )
    workbook = load_workbook(artifact.path)
    try:
        workbook["流水"]["E2"].fill = copy(workbook["流水"]["A2"].fill)
        workbook.save(artifact.path)
    finally:
        workbook.close()
    with pytest.raises(OutputVerificationError, match="颜色验证失败"):
        verify_output_workbook(
            artifact.path,
            preview,
            _targets(),
            _solution(),
            artifact.result_sheet,
        )
